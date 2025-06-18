import pandas as pd
from sqlalchemy import create_engine, text, Table, Column, Integer, String, Numeric, MetaData
import os
import unicodedata
import google.generativeai as genai
import re
import openpyxl
import json
from sqlalchemy.exc import IntegrityError
from agentic_sql_executor import safe_execute_agentic_sql
import difflib
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Access environment variables
db_host = os.getenv('DB_HOST')
db_port = os.getenv('DB_PORT')
db_name = os.getenv('DB_NAME')
db_user = os.getenv('DB_USER')
db_password = os.getenv('DB_PASSWORD')
google_api_key = os.getenv('GOOGLE_API_KEY')

# Database connection parameters
DB_PARAMS = {
    'dbname': db_name,
    'user': db_user,
    'password': db_password,
    'host': db_host,
    'port': db_port
}

# Configure Gemini API key
genai.configure(api_key=google_api_key)

def create_db_connection():
    """Create a database connection"""
    return create_engine(f"postgresql://{DB_PARAMS['user']}:{DB_PARAMS['password']}@{DB_PARAMS['host']}:{DB_PARAMS['port']}/{DB_PARAMS['dbname']}")

def drop_all_user_tables(engine):
    with engine.begin() as conn:
        result = conn.execute(text("""
            SELECT tablename FROM pg_tables
            WHERE schemaname = 'public';
        """))
        tables = [row[0] for row in result]
        for table in tables:
            print(f"[DEBUG] Dropping table: {table}")
            conn.execute(text(f"DROP TABLE IF EXISTS {table} CASCADE;"))

def sheet_matches_format(df):
    """Check if the DataFrame matches the expected format (columns and category headers)"""
    # Check for at least 14 columns
    if df.shape[1] < 14:
        return False
    # Check for expected columns in the first row (case-insensitive)
    expected = [
        "CONCEPTO", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE", "TOTAL"
    ]
    # Try to find a row that matches the expected columns
    for i in range(min(5, len(df))):
        row = df.iloc[i, :14].astype(str).str.upper().str.strip().tolist()
        if row == expected:
            return True
    # Or, just check that the first column has at least one all-caps string (category header)
    if df.iloc[:,0].apply(lambda x: isinstance(x, str) and x.strip().isupper()).any():
        return True
    return False

def safe_float(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def normalize_str(s):
    if not isinstance(s, str):
        return s
    s = s.strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s

def is_category_header(cell):
    if not isinstance(cell, str):
        return False
    norm = normalize_str(cell)
    # Remove all non-letter characters
    letters_only = re.sub(r'[^A-Z ]', '', norm)
    # Consider as header if at least 70% of the characters are uppercase letters or spaces and length > 3
    if len(letters_only) > 3 and (len(letters_only.replace(' ', '')) / max(1, len(norm.replace(' ', '')))) > 0.7:
        return True
    return False

def read_excel_sheet(sheet_name):
    """Read a specific Excel sheet and return a dictionary of DataFrames for each category"""
    excel_path = os.path.join('Excel', 'Flujo de Caja Grupo 11 2025 Abril.xlsx')
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)

    columns = [
        "CONCEPTO", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE", "TOTAL"
    ]
    month_cols = columns[1:13]

    categories = {}
    current_category = None
    current_data = []
    detected_categories = []
    for index, row in df.iterrows():
        first_col = row.iloc[0]
        print(f"[DEBUG] Row {index} first_col: '{first_col}' norm: '{normalize_str(first_col)}'")
        if is_category_header(first_col):
            # Always close the previous block, even if empty
            if current_category is not None:
                if current_data:  # Only add if there is at least one data row
                    cat_df = pd.DataFrame([r[:14] for r in current_data])
                    cat_df.columns = columns
                    for i, row2 in cat_df.iterrows():
                        total = sum([safe_float(row2[month]) for month in month_cols])
                        cat_df.at[i, 'TOTAL'] = total
                    cat_df = cat_df.drop_duplicates()
                    agg_row = [current_category.upper()]
                    for col in columns[1:]:
                        agg_row.append(cat_df[col].apply(safe_float).sum())
                    cat_df = pd.concat([
                        pd.DataFrame([agg_row], columns=columns),
                        cat_df
                    ], ignore_index=True)
                    categories[current_category] = cat_df
                    detected_categories.append(current_category)
                current_data = []
            current_category = normalize_str(first_col)
        else:
            # Skip empty rows
            if pd.isna(first_col) or (isinstance(first_col, str) and not first_col.strip()):
                continue
            if current_category:
                row_list = row.tolist()[:14]
                if len(row_list) < 14:
                    row_list += [None] * (14 - len(row_list))
                total = sum([safe_float(row_list[i]) for i in range(1, 13)])
                row_list[13] = total
                current_data.append(row_list)
    # Always close the last block
    if current_category is not None and current_data:
        cat_df = pd.DataFrame([r[:14] for r in current_data])
        cat_df.columns = columns
        for i, row2 in cat_df.iterrows():
            total = sum([safe_float(row2[month]) for month in month_cols])
            cat_df.at[i, 'TOTAL'] = total
        cat_df = cat_df.drop_duplicates()
        agg_row = [current_category.upper()]
        for col in columns[1:]:
            agg_row.append(cat_df[col].apply(safe_float).sum())
        cat_df = pd.concat([
            pd.DataFrame([agg_row], columns=columns),
            cat_df
        ], ignore_index=True)
        categories[current_category] = cat_df
        detected_categories.append(current_category)
    print(f"[DEBUG] Detected categories in sheet '{sheet_name}': {detected_categories}")
    return categories

def sanitize_table_name(name):
    # Replace invalid characters and ensure the name is a valid PostgreSQL identifier
    return (
        name.upper()
        .replace(' ', '_')
        .replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U')
        .replace('Ñ', 'N')
        .replace('?', '').replace('¿', '').replace('!', '').replace('¡', '')
        .replace('.', '').replace(',', '').replace('-', '_').replace('/', '_').replace(':', '_')
        .replace('__', '_')
    )

def truncate_name(name, suffix):
    max_len = 63 - len(suffix)
    return (name[:max_len] + suffix) if len(name) > max_len else (name + suffix)

def drop_old_tables(engine, old_prefix):
    with engine.begin() as conn:
        result = conn.execute(text(f"""
            SELECT tablename FROM pg_tables
            WHERE schemaname = 'public' AND tablename LIKE '{old_prefix}%';
        """))
        tables = [row[0] for row in result]
        for table in tables:
            print(f"[DEBUG] Dropping old table: {table}")
            conn.execute(text(f"DROP TABLE IF EXISTS {table} CASCADE;"))

def create_tables(categories):
    """Create PostgreSQL tables for each category"""
    engine = create_db_connection()
    # Drop old tables with the previous prefix
    drop_old_tables(engine, 'presupuesto_mercadeo_tanara_')
    table_info = []  # Store info for view creation

    # First pass: create and populate tables
    for category_key, df in categories.items():
        if '__' in category_key:
            sheet_name, category = category_key.split('__', 1)
        else:
            sheet_name, category = category_key, ''
        base_name = f"presupuesto_{sanitize_table_name(sheet_name)}_{sanitize_table_name(category)}"
        table_name = truncate_name(base_name, '_tbl')
        columns = ['CONCEPTO'] + [str(col) for col in df.columns[1:]]
        table_info.append((category, table_name, columns, base_name))
        print(f"[DEBUG] Creating table: {table_name}")
        with engine.begin() as conn:
            drop_table_sql = text(f"DROP TABLE IF EXISTS {table_name} CASCADE;")
            conn.execute(drop_table_sql)
            create_table_sql = f"""
            CREATE TABLE {table_name} (
                id SERIAL PRIMARY KEY,
                CONCEPTO VARCHAR(255),
                {', '.join([f'\"{col}\" DECIMAL(15,2)' for col in columns[1:]])}
            );
            """
            conn.execute(text(create_table_sql))
            print(f"[DEBUG] Table created: {table_name}")
            for _, row in df.iterrows():
                values = [row.iloc[0]] + [safe_float(row.iloc[i]) for i in range(1, len(columns))]
                placeholders = ', '.join([':val' + str(i) for i in range(len(values))])
                insert_sql = text(f"INSERT INTO {table_name} (CONCEPTO, {', '.join([f'\"{col}\"' for col in columns[1:]])}) VALUES ({placeholders})")
                params = {f'val{i}': v for i, v in enumerate(values)}
                conn.execute(insert_sql, params)

    # Second pass: create views
    for category, table_name, columns, base_name in table_info:
        view_name = truncate_name(base_name, '_vw')
        print(f"[DEBUG] Creating view for table: {table_name} as {view_name}")
        with engine.begin() as conn:
            drop_view_sql = text(f"DROP VIEW IF EXISTS {view_name} CASCADE;")
            conn.execute(drop_view_sql)
            create_view_sql = text(f"""
            CREATE VIEW {view_name} AS
            SELECT 
                '{category.upper()}' as categoria,
                {', '.join([f'SUM(\"{col}\") as \"{col}\"' for col in columns[1:]])}
            FROM {table_name};
            """)
            conn.execute(create_view_sql)
            print(f"[DEBUG] View created: {view_name} (from table: {table_name})")

    # Third pass: create unified view
    unified_view_name = truncate_name(f"vista_presupuesto_unificado", '_vw')
    print(f"[DEBUG] Creating unified view: {unified_view_name}")
    union_selects = []
    for category, table_name, columns, base_name in table_info:
        union_selects.append(f"SELECT CONCEPTO, {', '.join([f'\"{col}\"' for col in columns[1:]])} FROM {table_name} WHERE id = 1")
    gran_total_select = f"SELECT 'GRAN TOTAL' AS CONCEPTO, " + ', '.join([f'SUM("{col}") AS "{col}"' for col in columns[1:]]) + f" FROM (" + ' UNION ALL '.join([f"SELECT {', '.join([f'\"{col}\"' for col in columns[1:]])} FROM {table_name} WHERE id = 1" for _, table_name, columns, _ in table_info]) + ") t"
    unified_view_sql = f"""
    DROP VIEW IF EXISTS {unified_view_name} CASCADE;
    CREATE VIEW {unified_view_name} AS
    {(' UNION ALL\n').join(union_selects)}
    UNION ALL
    {gran_total_select};
    """
    with engine.begin() as conn:
        conn.execute(text(unified_view_sql))
        print(f"[DEBUG] Unified view created: {unified_view_name}")

def get_sheet_sample(excel_path, sheet_name, n_rows=10):
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    return df.head(n_rows).to_csv(index=False)

def ask_llm_for_table_def_gemini(sheet_csv, sheet_name):
    prompt = f"""
You are a data engineer. Given the following sample data from an Excel sheet named '{sheet_name}', suggest a PostgreSQL CREATE TABLE statement that best fits the data. Only output the SQL statement.

Sample data:
{sheet_csv}
"""
    model = genai.GenerativeModel("gemini-2.0-flash")
    response = model.generate_content(prompt)
    return response.text.strip()

def extract_tables_and_formulas(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    for sheet in wb.worksheets:
        print(f"\n--- Sheet: {sheet.title} ---")
        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                formula = cell.value if cell.data_type == 'f' else None
                row_data.append({
                    'coord': cell.coordinate,
                    'value': cell.value,
                    'formula': formula
                })
            print(row_data)

def extract_sheet_as_json(excel_path, sheet_name):
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    sheet = wb[sheet_name]
    data = []
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            row_data.append({
                "cell": cell.coordinate,
                "value": str(cell.value) if not isinstance(cell.value, (str, int, float, type(None))) else cell.value,
                "formula": str(cell.value) if cell.data_type == 'f' or not isinstance(cell.value, (str, int, float, type(None))) else (cell.value if cell.data_type == 'f' else None)
            })
        data.append(row_data)
    return data

def agentic_sql_generation(sheet_json, sheet_name):
    prompt = f"""
You are a data engineer agent. Given the following Excel sheet (as JSON, with cell values and formulas), analyze and identify all logical tables, columns, and relationships. For each table, generate a PostgreSQL CREATE TABLE or CREATE VIEW statement that replicates the logic, including formulas as SQL expressions or views. If you need more context, ask a clarifying question.

Sheet name: {sheet_name}
Sheet data (JSON):
{json.dumps(sheet_json, indent=2)}
"""
    model = genai.GenerativeModel("gemini-2.0-flash")
    response = model.generate_content(prompt)
    return response.text.strip()

def execute_sql(engine, sql):
    """Execute SQL statements in the database, only allowing CREATE TABLE statements for safety."""
    sql = sql.strip()
    statements = [stmt.strip() for stmt in sql.split(';') if stmt.strip()]
    with engine.begin() as conn:
        for stmt in statements:
            if stmt.upper().startswith('CREATE TABLE'):
                print(f"[DEBUG] Executing SQL:\n{stmt}\n")
                conn.execute(text(stmt))
            else:
                print(f"[WARNING] Skipping non-CREATE TABLE statement:\n{stmt}\n")

def load_excel_to_normalized_tables(engine, excel_path, sheet_name):
    """
    Loads data from the specified Excel sheet into the normalized tables:
    - activities
    - detailed_budget
    """
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
    print(df.head(10))
    # Find the start of the data (skip headers, project name, etc.)
    month_names = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
    header_row_idx = None
    for i, row in df.iterrows():
        row_upper = [str(cell).strip().upper() for cell in row]
        # Check if all month names and 'TOTAL' are present in the row (anywhere)
        if all(month in row_upper for month in month_names) and 'TOTAL' in row_upper:
            header_row_idx = i
            break
    if header_row_idx is None:
        print("[WARNING] Could not find header row. Preview:")
        print(df.head(10))
        # Optionally, prompt user for input:
        # header_row_idx = int(input("Enter the header row index (0-based): "))
        # Or fallback to first non-empty row:
        for i, row in df.iterrows():
            if any(str(cell).strip() for cell in row):
                header_row_idx = i
                break

    # Set up DataFrame with correct headers
    df2 = pd.read_excel(excel_path, sheet_name=sheet_name, header=header_row_idx)
    # Accept either 'CONCEPTO', 'ACTIVIDAD', or the first column as the activity name
    activity_col = None
    for col in df2.columns:
        if str(col).strip().upper() in ['CONCEPTO', 'ACTIVIDAD']:
            activity_col = col
            break
    if activity_col is None:
        activity_col = df2.columns[0]  # fallback to first column
    df2 = df2.dropna(subset=[activity_col])  # Drop rows with no activity name

    # Insert activities and detailed_budget
    with engine.begin() as conn:
        for _, row in df2.iterrows():
            activity = str(row[activity_col]).strip()
            if not activity or activity.upper() in ['TOTAL', 'GRAN TOTAL']:
                continue
            # Insert activity if not exists
            result = conn.execute(
                text("INSERT INTO activities (activity_name) VALUES (:activity_name) ON CONFLICT (activity_name) DO NOTHING RETURNING activity_id"),
                {"activity_name": activity}
            )
            if result.rowcount > 0:
                activity_id = result.fetchone()[0]
            else:
                # Get the existing activity_id
                activity_id = conn.execute(
                    text("SELECT activity_id FROM activities WHERE activity_name = :activity_name"),
                    {"activity_name": activity}
                ).scalar()
            # Insert monthly values
            for month in month_names:
                value = row.get(month, None)
                if pd.isna(value):
                    continue
                conn.execute(
                    text("""
                        INSERT INTO detailed_budget (activity_id, month, amount)
                        VALUES (:activity_id, :month, :amount)
                    """),
                    {"activity_id": activity_id, "month": month, "amount": float(value)}
                )
    print("[INFO] Data loaded into activities and detailed_budget tables.")

def create_normalized_tables(engine):
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS activities (
                activity_id SERIAL PRIMARY KEY,
                activity_name VARCHAR(255) UNIQUE NOT NULL
            );
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS detailed_budget (
                budget_id SERIAL PRIMARY KEY,
                activity_id INTEGER REFERENCES activities(activity_id),
                month VARCHAR(20),
                amount NUMERIC(15,2)
            );
        """))

def parse_create_table_columns(sql):
    """
    Parse CREATE TABLE statement to extract table name and column names.
    Returns (table_name, [col1, col2, ...])
    """
    match = re.search(r'CREATE TABLE\s+([^\s(]+)\s*\((.*?)\);', sql, re.DOTALL | re.IGNORECASE)
    if not match:
        print("[ERROR] Could not parse CREATE TABLE statement.")
        return None, []
    table_name = match.group(1)
    columns_block = match.group(2)
    # Remove PRIMARY KEY, etc.
    columns = []
    for line in columns_block.splitlines():
        line = line.strip().strip(',')
        if not line or line.upper().startswith('PRIMARY KEY'):
            continue
        col = line.split()[0].strip('"')
        columns.append(col)
    return table_name, columns

def find_header_row(df, expected_cols):
    for i in range(min(10, len(df))):
        row = [str(cell).strip().upper() for cell in df.iloc[i]]
        matches = sum(1 for col in expected_cols if col in row)
        if matches >= max(2, len(expected_cols)//2):  # at least half match
            return i, row
    return None, None

def find_header_row_fuzzy(df, expected_cols, min_matches=0.5):
    """
    Try to find the header row by fuzzy matching expected columns.
    Returns (header_row_index, matched_columns) or (None, None) if not found.
    """
    expected_set = set([col.strip().upper() for col in expected_cols])
    for i in range(min(15, len(df))):
        row = [str(cell).strip().upper() for cell in df.iloc[i]]
        matches = sum(
            1 for col in row
            if difflib.get_close_matches(col, expected_set, n=1, cutoff=0.7)
        )
        if matches >= max(2, int(len(expected_cols) * min_matches)):
            return i, row
    return None, None

def insert_excel_data_to_table(engine, excel_path, sheet_name, create_table_sql):
    table_name, columns = parse_create_table_columns(create_table_sql)
    if not table_name or not columns:
        print("[ERROR] Table name or columns not found.")
        return

    # Check if table exists before inserting
    with engine.connect() as conn:
        table_exists = conn.execute(
            text("SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_name = :table_name);"),
            {"table_name": table_name.lower()}
        ).scalar()
        if not table_exists:
            print(f"[ERROR] Table {table_name} does not exist in the database. Skipping insert.")
            return

    # Remove 'id' if it's the first column and likely SERIAL
    if columns and columns[0].lower() == 'id':
        columns = columns[1:]

    df_raw = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)

    # Special header detection for 'Costo alquiler equipo y comb'
    header_row_idx = None
    if sheet_name == "Costo alquiler equipo y comb":
        for i, row in df_raw.iterrows():
            row_str = [str(cell).strip().lower() if not pd.isna(cell) else '' for cell in row]
            if 'equipo' in row_str:
                header_row_idx = i
                print(f"[HEADER DETECTION] Using row {header_row_idx} as header for '{sheet_name}': {row_str}")
                break
    else:
        header_row_idx, header_row = find_header_row_fuzzy(df_raw, columns)
        if header_row_idx is None:
            # Fallback: use first non-empty row as header
            for i, row in df_raw.iterrows():
                if any(str(cell).strip() for cell in row):
                    header_row_idx = i
                    header_row = [str(cell).strip() for cell in row]
                    print(f"[FALLBACK] Using row {header_row_idx} as header: {header_row}")
                    break

    if header_row_idx is not None:
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header_row_idx)
        # Data mapping: if SQL columns are generic, map by position; else, map by name
        if all(col.lower().startswith('unnamed') or col.lower().startswith('column') for col in columns):
            mapping = list(df.columns[:len(columns)])
            print(f"[MAPPING] Using positional mapping for generic columns: {mapping}")
        else:
            mapping = []
            for col in columns:
                if col in df.columns:
                    mapping.append(col)
                else:
                    mapping.append(None)
            print(f"[MAPPING] Using name-based mapping: {mapping}")
    else:
        print(f"[WARNING] No header found for table {table_name} in sheet {sheet_name}. Inserting by position.")
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
        df = df.iloc[:, :len(columns)]
        mapping = list(df.columns)
        print(f"[MAPPING] Using positional mapping (no header): {mapping}")

    print(f"[DEBUG] DataFrame preview:\n{df.head()}")
    print(f"[DEBUG] SQL columns: {columns}")
    print(f"[DEBUG] Excel->SQL mapping: {mapping}")

    with engine.begin() as conn:
        for _, row in df.iterrows():
            # Skip rows that are all-NaN or look like headers/footers
            values = [row[col] if col is not None and col in row else None for col in mapping]
            if all(v is None or (isinstance(v, float) and pd.isna(v)) for v in values):
                continue
            if all(isinstance(v, str) for v in values if v is not None) and len(values) > 0:
                # Looks like a header/footer row
                continue
            if any(isinstance(v, str) and 'TOTAL' in v.upper() for v in values if v is not None):
                continue
            placeholders = ', '.join([f':val{i}' for i in range(len(values))])
            col_names = ', '.join([f'"{c}"' for c in columns])
            sql = text(f'INSERT INTO {table_name} ({col_names}) VALUES ({placeholders})')
            params = {f'val{i}': v for i, v in enumerate(values)}
            try:
                conn.execute(sql, params)
            except Exception as e:
                print(f"[ERROR] Failed to insert row: {e}")
    print(f"[INFO] Inserted data into {table_name} from {sheet_name}")

def split_sheet_into_tables(df, min_rows=3):
    """
    Splits a DataFrame into a list of DataFrames, each representing a table.
    Tables are separated by blank rows (all NaN or empty).
    """
    tables = []
    current = []
    for idx, row in df.iterrows():
        if row.isnull().all() or all(str(cell).strip() == '' or pd.isna(cell) for cell in row):
            if len(current) >= min_rows:
                tables.append(pd.DataFrame(current, columns=df.columns))
            current = []
        else:
            current.append(row)
    if len(current) >= min_rows:
        tables.append(pd.DataFrame(current, columns=df.columns))
    return tables

def prototype_agentic_table_extraction_and_sql(excel_path):
    """
    For each sheet in the Excel file:
      - Detect blocks by finding a row with a non-empty string (block title), followed by a header row, then data rows.
      - Use the block title and sheet name to generate the table name (sanitize for SQL).
      - Use the header row for columns, and extract data rows until a row with 'Gran total' or a blank row.
      - Exclude the 'Gran total' row from data insertion.
      - Pass the correct table name and data to the LLM prompt and data insertion logic.
    """
    import re
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    engine = create_db_connection()
    for sheet in wb.worksheets:
        sheet_name = sheet.title
        if sheet_name != "Costo alquiler equipo y comb":
            continue
        df_raw = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
        n_rows, n_cols = df_raw.shape
        i = 0
        while i < n_rows:
            # Find block title row (non-empty string, not a header)
            row = df_raw.iloc[i]
            if any(isinstance(cell, str) and cell.strip() and not cell.strip().lower() in ["equipo", "qty", "precio", "meses", "total"] for cell in row):
                block_title = next((str(cell).strip() for cell in row if isinstance(cell, str) and cell.strip()), None)
                # Next row is header
                if i+1 < n_rows:
                    header_row = df_raw.iloc[i+1]
                    headers = [str(cell).strip().lower() for cell in header_row]
                    # Find data rows
                    data_rows = []
                    for j in range(i+2, n_rows):
                        data_row = df_raw.iloc[j]
                        # Stop at 'Gran total' or blank row
                        if any(isinstance(cell, str) and 'gran total' in cell.lower() for cell in data_row) or data_row.isnull().all():
                            break
                        data_rows.append([cell for cell in data_row])
                    if data_rows:
                        # Sanitize table name
                        table_name = f"presupuesto_{sheet_name.lower().replace(' ', '_')}_{block_title.lower().replace(' ', '_')}"
                        table_name = re.sub(r'[^a-zA-Z0-9_]', '', table_name)
                        # Prepare DataFrame
                        df_block = pd.DataFrame(data_rows, columns=headers)
                        # Remove empty columns
                        df_block = df_block.loc[:, ~df_block.columns.str.contains('^unnamed', case=False)]
                        # LLM prompt
                        data = df_block.to_dict(orient='records')
                        prompt = f"""
You are a data engineer agent. Given the following table, generate a valid PostgreSQL CREATE TABLE statement named '{table_name}' with appropriate column types. Only output the SQL statement.\nTable data (JSON):\n{data}
"""
                        print(f"[AGENTIC] Processing block '{block_title}' in sheet: {sheet_name}")
                        model = genai.GenerativeModel("gemini-2.0-flash")
                        response = model.generate_content(prompt)
                        sql = response.text.strip()
                        print(f"[AGENTIC LLM SQL OUTPUT for {sheet_name} Block {block_title}]:\n{sql}\n")
                        safe_execute_agentic_sql(engine, sql)
                        # Insert data
                        if 'CREATE TABLE' in sql:
                            create_table_sqls = re.findall(r'CREATE TABLE[\s\S]*?;', sql, re.IGNORECASE)
                            for create_table_sql in create_table_sqls:
                                # Use our DataFrame for insertion
                                # Patch: call insert_excel_data_to_table with our DataFrame
                                # We'll use a custom version for this block
                                table_name_from_sql, columns_from_sql = parse_create_table_columns(create_table_sql)
                                if not table_name_from_sql or not columns_from_sql:
                                    continue
                                with engine.begin() as conn:
                                    for _, row in df_block.iterrows():
                                        values = [row.get(col, None) for col in columns_from_sql]
                                        if all(v is None or (isinstance(v, float) and pd.isna(v)) for v in values):
                                            continue
                                        placeholders = ', '.join([f':val{i}' for i in range(len(values))])
                                        col_names = ', '.join([f'"{c}"' for c in columns_from_sql])
                                        sql_insert = text(f'INSERT INTO {table_name_from_sql} ({col_names}) VALUES ({placeholders})')
                                        params = {f'val{i}': v for i, v in enumerate(values)}
                                        try:
                                            conn.execute(sql_insert, params)
                                        except Exception as e:
                                            print(f"[ERROR] Failed to insert row: {e}")
                                print(f"[INFO] Inserted data into {table_name_from_sql} from block '{block_title}' in sheet '{sheet_name}'")
                    i = j  # Skip to after this block
                else:
                    i += 1
            else:
                i += 1

def export_costo_directo(engine, excel_path):
    sheet_name = 'Materiales por comprar'
    df = pd.read_excel(excel_path, sheet_name=sheet_name, usecols='B:E', skiprows=5, nrows=11)
    df.columns = df.iloc[0]  # Set the first row as column titles
    df = df.drop(0)  # Remove the row used for column titles
    table_name = 'costo_directo'
    print(f"[INFO] Exporting '{sheet_name}' to table '{table_name}'")
    df.to_sql(table_name, engine, if_exists='replace', index=False)

    print(f"[INFO] Exported '{sheet_name}' to PostgreSQL table '{table_name}'")



def main():
    print("Reading Excel file...")
    engine = create_db_connection()
    drop_all_user_tables(engine)  # Drop all user tables before proceeding
    create_normalized_tables(engine)  # Ensure normalized tables exist
    excel_path = os.path.join('Excel', 'Flujo de Caja Grupo 11 2025 Abril.xlsx')
    xls = pd.ExcelFile(excel_path)
    processed_sheets = []
    all_categories = {}
    for sheet_name in xls.sheet_names:
        print(f"Checking sheet: {sheet_name}")
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
        if sheet_matches_format(df):
            print(f"[INFO] Sheet '{sheet_name}' matches format. Processing...")
            categories = read_excel_sheet(sheet_name)
            for k, v in categories.items():
                all_categories[f"{sheet_name}::{k}"] = v
            processed_sheets.append(sheet_name)
        else:
            print(f"[INFO] Sheet '{sheet_name}' does not match format. Skipping.")
    if all_categories:
        print("Creating database tables...")
        create_tables(all_categories)
    print("Process completed!")
    print("Sheets processed:", processed_sheets)

    # LLM SQL generation and execution (Gemini)
    for sheet_name in processed_sheets:
        print(f"\n--- {sheet_name} ---")
        sample_csv = get_sheet_sample(excel_path, sheet_name)
        sql = ask_llm_for_table_def_gemini(sample_csv, sheet_name)
        print("[LLM SQL OUTPUT]\n", sql)
        execute_sql(engine, sql)

    # Now run the LLM prototype for the target sheet after all cleanup and table creation
    print("[INFO] Running LLM prototype for 'Costo alquiler equipo y comb' after all cleanup...")
    prototype_agentic_table_extraction_and_sql(excel_path)

    # Uncomment the following lines to run the agentic workflow on a single sheet
    #excel_path = "Excel/Flujo de Caja Grupo 11 2025 Abril.xlsx"
    #sheet_name = "Presupuesto Mercadeo Chepo"  # or any sheet you want
    #sheet_json = extract_sheet_as_json(excel_path, sheet_name)
    #sql = agentic_sql_generation(sheet_json, sheet_name)
    #print("[AGENTIC LLM SQL OUTPUT]\n", sql)

    clarifications = """
    1. Data Types: Use NUMERIC for all numbers, VARCHAR(255) for text.
    2. Monthly Data: Store as multiple columns per activity.
    3. Project/Row Totals: Calculate on the fly with SQL.
    4. Missing Data: Store as NULL.
    5. Project Name: Store in a separate table and reference in each row.
    """
    followup_prompt = sql + "\n\nHere are the answers to your questions:\n" + clarifications + "\nNow, please generate the PostgreSQL CREATE TABLE and CREATE VIEW statements."
    # Send followup_prompt to Gemini and print the result
    model = genai.GenerativeModel("gemini-2.0-flash")
    response = model.generate_content(followup_prompt)
    print("[AGENTIC LLM FINAL SQL OUTPUT]\n", response.text.strip())

    # Load data into normalized tables
    load_excel_to_normalized_tables(engine, excel_path, sheet_name)

if __name__ == "__main__":
    # Uncomment the following line to run the extraction prototype
    # extract_tables_and_formulas("Excel/Flujo de Caja Grupo 11 2025 Abril.xlsx")
    main() 